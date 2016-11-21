# coding=utf-8

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

def fill2sheet(sheet, sheet_name, records, headers=None):
    sheet.title=sheet_name
    if headers:
        sheet.append(headers)
    for record in records:
        sheet.append(record)

def write2xlsx(filename, records, headers=[], sheets=[]):
    wb = Workbook()

    if sheets and isinstance(sheets, list):
        for index in range(len(sheets)):
            ws = wb.create_sheet()
            record = records[index] if index < len(records) else []
            header = headers[index] if index < len(headers) else []
            fill2sheet(ws, sheets[index], record, header)
    elif sheets and isinstance(sheets, str):
        ws = wb.create_sheet()
        fill2sheet(ws, sheets, records, headers)
    wb.save(filename)

def readxlsx(filename, sheet_name=None, startrow=0, endrow=None):
    wb = load_workbook(filename)
    if sheet_name:
        ws = wb.get_sheet_by_name(sheet_name)
    else:
        sheetnames = wb.get_sheet_names()  
        ws = wb.get_sheet_by_name(sheetnames[0])
    line_no = -1
    records = []
    for row in ws.rows:
        line_no += 1
        if endrow is not None and line_no >= endrow:
            break
        if startrow is not None and line_no < startrow:
            continue

        data = []
        for cell in row:
            data.append(cell.value)
        records.append(data)
    return records

if __name__ == '__main__':
    records = []
    for i in range(4):
        record = []
        for k in range(100):
            record.append(range((k+i)*10, (k+i)*100,))
        records.append(record)
    headers = []
    sheets = ['1', '2', '3', '4', '5']
    #write2xlsx("test.xlsx", records, headers, sheets)
    records = readxlsx("test.xlsx", "1", 1)
    for record in records:
        print ','.join(record)

